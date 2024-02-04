# -*- coding:utf-8 -*-
import os
import time
import traceback

import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
import re

from openpyxl.comments import Comment
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, GradientFill
from openpyxl.styles import Font
from openpyxl.styles import Border, Side

COOKIE_STR = 'experimentation_subject_id=IjFjYjNlNTYzLWRiZTAtNDM3ZC05MzFlLTk4OGY3MjZkYmI1MSI%3D--642f87fac21ad96137b0ca2b46f781df1d1781bf; tradeMA=68; _snvd=1702028146357WryCRiG+g6C; idsLoginUserIdLastTime=20013921; isso_ld=true; isso_us=20013921; authId=siCA726D56F2D41562D02F8BA81A3AF3F1; secureToken=B23604D83A0200059C3CEDE2769F9019; JSESSIONID=82lxflvdbvfq27ugozcp0jv'
URL_GET_TASK_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTaskInfoByKind'
RUL_GET_TPS_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTpsInfo'
LINK_TYPE = "-INC-"
PAGE_SIZE = 100
EXCEL_FILE = 'D:\\file\\202402\\rdrs_links.xlsx'


def cookie_to_dict():
    """
    将字符串形式的cookie转为字典格式
    :return:
    """
    # 列表推导
    cookie_entries = [(entry.strip().split('=')[0], entry.strip().split('=')[1]) for entry in COOKIE_STR.split(';')]
    # 字典推导
    cookie_dict = {k: v for k, v in cookie_entries}
    return cookie_dict


cookies = cookie_to_dict()
headers = {'Content-Type': 'application/json'}
link_page_data_param = {
    'taskNo': '',
    'kind': LINK_TYPE,
    'taskName': '',
    'status': '',
    'system': '',
    'page': 1,
    'size': 10
}


def get_page_link_data(page, size):
    """
    获取分页数据
    :param page: 页号
    :param size: 每页数据条数
    :return: 该页数据
    """
    link_page_data_param['page'] = page
    link_page_data_param['size'] = size
    # 在Cookie Version 0中规定空格、方括号、圆括号、等于号、逗号、双引号、斜杠、问号、@，冒号，分号等特殊符号都不能作为Cookie的内容。
    response = requests.get(URL_GET_TASK_INFO, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        return None
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        return None
    result = response_dict.get('result')
    data = result.get('data')
    if data is None or len(data) == 0:
        return None
    total = result.get('total')
    size = result.get('size')
    return data, total, size


def get_tps_info_0(link_info):
    """
    获取某条链路指定的tps信息
    :param link_info: 链路信息
    :return:
    """
    param = {'taskName': link_info.get('taskName')}
    response = requests.get(RUL_GET_TPS_INFO, cookies=cookies, headers=headers, params=param)
    if response.status_code != 200:
        link_info['haveData'] = True
        link_info['tps'] = str(['9999' for _ in range(15)])
        return
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        link_info['haveData'] = True
        link_info['tps'] = str(['9999' for _ in range(15)])
        return
    # 只取最近15分钟这一项
    fifteen_tps_dict = response_dict.get('result')[2]
    # value ["7847", "5509", "6853", "5951", "5517", "4141", "3960", "4155", "4476", "5328", "6633", "7318",…]
    tps = fifteen_tps_dict.get('value')
    link_info['haveData'] = tps.count('0') != 15
    link_info['tps'] = str(tps)


def set_tps_info(link_list):
    """
    获取链路对应的tps信息
    :param link_list: 链路列表
    :return:
    """
    executor = ThreadPoolExecutor(max_workers=5)
    executor.map(get_tps_info_0, link_list)
    # 这里等待所有任务执行完毕再返回这个函数
    executor.shutdown()


def get_link_list():
    """
    获取链路列表
    :return:
    """
    link_list = []
    # 先探测一下数据总条数
    page_data = get_page_link_data(1, 10)
    if page_data is None:
        return link_list
    total = page_data[1]
    print('链路类型: {}'.format('增量' if LINK_TYPE.find('INC') != -1 else '全量'))
    print('链路总量: {}'.format(total))
    print('每页数量: {}'.format(PAGE_SIZE))
    print('最大页数: {}'.format(int(total / PAGE_SIZE) + 2))
    print('开始获取页数据...')
    # 下面的代码和  for (int i = 1; i <= total / PAGE_SIZE + 1; i++)  是等价的
    for page_no in range(1, int(total / PAGE_SIZE) + 2):
        page_data = get_page_link_data(page_no, PAGE_SIZE)
        if page_data is None:
            continue
        page_link_list = page_data[0]
        set_tps_info(page_link_list)
        link_list.extend(page_link_list)
        print('当前页号: {} 数据获取完成'.format(page_no))
    print('所有页数据获取完成！')
    return link_list


def ftime(f=''):
    if f == '-':
        return time.strftime('%Y-%m-%d  %H:%M:%S', time.localtime())
    if f == '':
        time_stamp = time.time()
        ends = str(time_stamp - int(time_stamp))[2:4]
        return time.strftime('%Y%m%d%H%M%S', time.localtime()) + ends
    if f == '_':
        return time.strftime('%Y{}%m{}%d{}',
                             time.localtime()).format("年", "月", "日")


def is_my_sheet_name(sheet_name):
    if sheet_name is None:
        return False
    pattern = r'^\d{4}年\d{2}月\d{2}日-(\d+)$'
    match = re.match(pattern, sheet_name.strip())
    if match:
        return True
    else:
        return False


def my_sheet_name_filter(sheet_names):
    """
    过滤出符合要求的sheet_name列表
    :param sheet_names:
    :return:
    """
    if sheet_names is None or len(sheet_names) == 0:
        return []
    my_sheet_names = [sheet_name.strip() for sheet_name in sheet_names if is_my_sheet_name(sheet_name)]
    my_sheet_names.sort()
    return my_sheet_names


def write_link_list(link_list):
    """
    写入link_list到excel中
    :param link_list: 链路列表
    :return: 工作表
    """
    wb = Workbook()
    if os.path.exists(EXCEL_FILE):
        # 存在则使用已存在的文件
        wb = openpyxl.load_workbook(EXCEL_FILE)
    # 获取全局最大序号
    max_no = 0
    my_sheet_names = my_sheet_name_filter(wb.sheetnames)
    if my_sheet_names is None:
        raise Exception("没有符合指定模式的sheet页，请检查")
    if len(my_sheet_names) != 0:
        max_no = max(my_sheet_names).split('-')[-1]
    max_no = int(max_no) + 1
    sheet_name = "{}-{}".format(ftime('_'), max_no)
    ws = wb.create_sheet(sheet_name)
    print('创建sheet页: {}'.format(sheet_name))

    header = ['taskId', 'taskName', 'taskStatus', 'systemName', 'clusterName', 'workIp', 'startTime', 'maxBinlogFile',
              'nowBinlogFile', 'minBinlogFile', 'nowPosition', 'trace', 'delay', 'delayOfnum', 'haveData', 'tps']
    ws.append(header)
    for link_dict in link_list:
        row = [link_dict.get(key) for key in header]
        ws.append(row)
        print('链路写入成功: {} !'.format(link_dict.get('taskName')))
    wb.save(EXCEL_FILE)
    print('写入 {} 条链路数据到excel成功！'.format(len(link_list)))
    return wb


def get_last_two_sheet(wb):
    """
    获取sheet名排序中最后两个sheet
    :param wb:
    :return:
    """
    my_sheet_names = my_sheet_name_filter(wb.sheetnames)
    if len(my_sheet_names) < 1:
        raise Exception("没有符合指定模式的sheet页")
    if len(my_sheet_names) == 1:
        return wb[my_sheet_names[-1]], wb[my_sheet_names[-1]]
    last_one = wb[my_sheet_names[-1]]
    last_two = wb[my_sheet_names[-2]]
    return last_one, last_two


last_one_task_map_row = {}
last_two_task_map_row = {}


def construct_task_map_row(wb):
    """
    构建任务号到行的索引
    :return: None
    """
    last_one, last_two = get_last_two_sheet(wb)
    global last_one_task_map_row
    # 字典推导构建任务号到行的索引，任务号作为key，行作为value，这样就可以通过任务号快速定位到行。排除空行
    last_one_task_map_row = {str(row[0].value).strip(): row for row in last_one.rows if row[0].value is not None and len(str(row[0].value).strip()) != 0}
    global last_two_task_map_row
    # 字典推导构建任务号到行的索引，任务号作为key，行作为value，这样就可以通过任务号快速定位到行。排除空行
    last_two_task_map_row = {str(row[0].value).strip(): row for row in last_two.rows if row[0].value is not None and len(str(row[0].value).strip()) != 0}


def status_change_marker(last_one, last_two):
    """
    在最后一个sheet页中标记状态变化，有变化的链路标记为 EBE766 黄色
    :param last_one: 倒数第一个sheet页
    :param last_two: 倒数第二个sheet页
    :return: None
    """
    fill = PatternFill(start_color='EBE766', end_color='EBE766', fill_type='solid')
    font = Font(name='微软雅黑', size=11, bold=True, italic=False, color='E0324C')
    # 构建一个线条样式 粗线，棕色
    side = Side(style='thick', color='590C00')
    # 构建一个边框样式
    border = Border(left=side, top=side, right=side, bottom=side)
    last_one_keys = last_one_task_map_row.keys()
    for task_id in last_one_keys:
        # 倒数第一sheet页中task_id对应的行
        last_one_cells = last_one_task_map_row.get(task_id)
        # 倒数第二sheet页中task_id对应的行
        last_two_cells = last_two_task_map_row.get(task_id)
        if last_two_cells is None:
            # 为职责清晰，这里不顺便标记last_one sheet页中多出来的链路，而是由下面的 link_num_change_marker 进行标记
            continue
        # row是一个元组，只能通过下标取链路状态值
        last_one_task_status = last_one_cells[2]
        last_two_task_status = last_two_cells[2]
        if last_one_task_status.value.strip().upper() == last_two_task_status.value.strip().upper():
            continue
        # 状态有变化，则标记颜色
        last_one_task_status.font = font
        last_one_task_status.border = border
        comment = Comment("之前的状态是: " + last_two_task_status.value.strip().upper(), "雷建忠(20013921)")
        last_one_task_status.comment = comment
        for cell in last_one_cells:
            cell.fill = fill


def link_num_change_marker(last_one, last_two):
    """
    在最后一个sheet页中标记链路数量的变化，多出来的链路标记为灰绿色 80B492 ，少了的链路追加到sheet页最后，标记为红色 E0324C
    :param last_one: 倒数第一个sheet页
    :param last_two: 倒数第二个sheet页
    :return: None
    """
    # 多出来的链路标记为灰绿色
    fill_increased = PatternFill(start_color='80B492', end_color='80B492', fill_type='solid')
    # 少了的链路标记为红色，并追加到sheet页最后
    fill_decreased = PatternFill(start_color='E0324C', end_color='E0324C', fill_type='solid')
    # 获取倒数第一个sheet的任务号集合，但是排除掉那些填充颜色为红色的任务，因为红色的任务是该sheet页和前一个sheet页比较以后新增加的任务，不是其本身的任务
    last_one_task_ids = {task_id for task_id, row in last_one_task_map_row.items() if row[0].fill is None or row[0].fill.start_color.value != '00E0324C'}
    # 获取倒数第二sheet的任务号集合，但是排除掉那些填充颜色为红色的任务，因为红色的任务是该sheet页和前一个sheet页比较以后新增加的任务，不是其本身的任务
    last_two_task_ids = {task_id for task_id, row in last_two_task_map_row.items() if row[0].fill is None or row[0].fill.start_color.value != '00E0324C'}
    # 获取 last_one_task_ids 中有的，但是last_two_task_ids中没有的任务，即新增加的任务
    increased_task_ids = last_one_task_ids - last_two_task_ids
    print("increased_task_ids: ", increased_task_ids)
    if len(increased_task_ids) > 0:
        for task_id in increased_task_ids:
            task_row = last_one_task_map_row.get(task_id)
            if task_row is None:
                continue
            for cell in task_row:
                cell.fill = fill_increased

    # 获取 last_two_task_ids 中有的，但是 last_two_task_ids 中没有的任务，即减少的任务
    decreased_task_ids = last_two_task_ids - last_one_task_ids
    print("decreased_task_ids: ", decreased_task_ids)
    decreased_tasks = []
    if len(decreased_task_ids) > 0:
        for task_id in decreased_task_ids:
            task_row = last_two_task_map_row.get(task_id)
            if task_row is None:
                continue
            decreased_tasks.append(task_row)
    if len(decreased_tasks) > 0:
        # 在最后一行的下方插入一个空行
        last_one.insert_rows(last_one.max_row + 1)
        # 获取到减少的链路列表之后，将这些链路插入到最后一个sheet页中
        max_row_num = last_one.max_row + 2
        for decreased_task in decreased_tasks:
            for cell in decreased_task:
                new_cell = last_one.cell(row=max_row_num, column=cell.column, value=cell.value)
                new_cell.fill = fill_decreased
            max_row_num += 1


func_change_markers = [globals()[name] for name in globals() if name.endswith('_change_marker') and name != 'func_change_markers']


def mark_link_change(wb):
    """
    标记链路变化
    :param wb:
    :return:
    """
    construct_task_map_row(wb)
    last_one, last_two = get_last_two_sheet(wb)
    for func_change_marker in func_change_markers:
        try:
            func_change_marker(last_one, last_two)
        except Exception as e:
            traceback.print_exc()
    wb.save(EXCEL_FILE)


def main():
    # 获取全量的链路列表
    link_list = get_link_list()
    # 写入链路列表到excel中
    wb = write_link_list(link_list)
    # 标记链路变化
    mark_link_change(wb)


if __name__ == '__main__':
    main()


