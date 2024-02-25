# -*- coding:utf-8 -*-
import os
import re
import sys
from bs4 import BeautifulSoup
import requests
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.workbook import Workbook
import pymysql.cursors
import time

"""
爬虫的方式从itsm获取vip信息
"""

COOKIE_STR = 'JSESSIONID=dboFhBThvQ2UtmXhSkKA7vEG.itsmprdapp208; experimentation_subject_id=IjFjYjNlNTYzLWRiZTAtNDM3ZC05MzFlLTk4OGY3MjZkYmI1MSI%3D--642f87fac21ad96137b0ca2b46f781df1d1781bf; tradeMA=68; idsLoginUserIdLastTime=20013921; route=037ffa392b986cb6b73b7afb4792e5e2; SN_SESSION_ID=a6a43a18-e657-4a2e-ab3d-1ab6b98d7574; isso_ld=true; isso_us=20013921; scm-prd=siC07E171F39F366D7AD3261D61FEAE5EA; custno=20013921; _snma=1%7C170702760811448759%7C1707027608114%7C1708581444417%7C1708584359850%7C7%7C4; _snmp=17085843597737784; _snmb=170858435985351870%7C1708584359863%7C1708584359854%7C1; _snvd=1708581338362qBSqLmb+EpD; secureToken=C4CC3AC061BB215FB7DE014850E1E3B0; authId=si0BE28BF74BA18D5169B3C08ABC583A06'
URL_GET_TASK_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTaskInfoByKind'
URL_GET_TPS_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTpsInfo'
URL_GET_CONFIG_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getConfigInfo'
# -INC-  -SINK-
LINK_TYPE = "-INC-"
PAGE_SIZE = 100
EXCEL_FILE = 'D:\\file\\202402\\vip_info.xlsx'
URL_GET_IP_INFO = 'http://itsm.cnsuning.com/traffic-web-in/api/getServerByIpInfo.htm'


def cookie_to_dict():
    """
    将字符串形式的cookie转为字典格式
    :return:
    """
    # 列表推导
    cookie_entries = [(entry.strip().split('=')[0], entry.strip().split('=')[1]) for entry in COOKIE_STR.split(';')]
    # 字典推导
    cookie_dict = {k: v for k, v in cookie_entries}
    return cookie_dict


cookies = cookie_to_dict()
link_page_data_param = {
    'taskNo': '',
    'kind': LINK_TYPE,
    'taskName': '',
    'status': '',
    'system': '',
    'page': 1,
    'size': 10
}


def get_page_link_data(page, size):
    """
    获取分页数据
    :param page: 页号
    :param size: 每页数据条数
    :return: 该页数据
    """
    link_page_data_param['page'] = page
    link_page_data_param['size'] = size
    headers = {'Content-Type': 'application/json'}
    # 在Cookie Version 0中规定空格、方括号、圆括号、等于号、逗号、双引号、斜杠、问号、@，冒号，分号等特殊符号都不能作为Cookie的内容。
    response = requests.get(URL_GET_TASK_INFO, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        return None
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        return None
    result = response_dict.get('result')
    data = result.get('data')
    if data is None or len(data) == 0:
        return None
    total = result.get('total')
    size = result.get('size')
    return data, total, size


def get_link_list():
    """
    获取链路列表
    :return:
    """
    link_list = []
    # 先探测一下数据总条数
    page_data = get_page_link_data(1, 10)
    if page_data is None:
        return link_list
    total = page_data[1]
    print('链路类型: {}'.format('增量' if LINK_TYPE.find('INC') != -1 else '全量'))
    print('链路总量: {}'.format(total))
    print('每页数量: {}'.format(PAGE_SIZE))
    print('最大页数: {}'.format(int(total / PAGE_SIZE) + 2))
    print('开始获取页数据...')
    # 下面的代码和  for (int i = 1; i <= total / PAGE_SIZE + 1; i++)  是等价的
    for page_no in range(1, int(total / PAGE_SIZE) + 2):
        page_data = get_page_link_data(page_no, PAGE_SIZE)
        if page_data is None:
            continue
        page_link_list = page_data[0]
        link_list.extend(page_link_list)
        print('当前页号: {} 数据获取完成'.format(page_no))
    print('所有页数据获取完成！')
    return link_list


def get_vip_from_task_name(task_name):
    """
    从taskName中获取vip
    :return:
    """
    # 找到上面字符串中包含的ip
    return re.findall(r"\d+.\d+.\d+.\d+", task_name)[0]


def get_bk_ip(task_name, db_ip):
    """
    从itsm查备份网络ip
    :param task_name:
    :param db_ip:
    :return:
    """
    # post方法请求URL_GET_IP_INFO接口
    post_param = {
        "systemEnName": "SNRS",
        "key": "a4d9bef0270946c7b2b53999eb1488fa",
        "authType": 0,
        "timeStamp": int(round(time.time() * 1000)),
        "ip": str(db_ip).strip()
    }
    headers = {'Content-Type': 'application/json'}
    response = requests.post(URL_GET_IP_INFO, cookies=None, headers=headers, json=post_param)
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        # 特殊写法，解决并发打印错乱问题
        print(f'{task_name} {db_ip} 请求URL_GET_IP_INFO接口失败！\n', end='')
        return 'NA'
    datas = response_dict.get('datas')
    if len(datas) == 0:
        print(f'{task_name} URL_GET_IP_INFO未返回该ip信息: NA \n', end='')
        return 'NA'
    bk_ip = str(datas[0].get('backupnetworkIP')).strip()
    if bk_ip == '':
        print(f'{task_name} 返回bkip为空:NA \n', end='')
        return 'NA'
    return bk_ip


def query_ip_info(task_name, ip):
    """
    获取内网ip的基本信息
    :param task_name:
    :param ip:
    :return:
    """
    if ip is None or ip == '':
        return False, '传入ip为空'

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0',
        'Cookie': COOKIE_STR
    }
    url = f"http://itsm.cnsuning.com/traffic-web-in/intranetip/loadIntranetipList.htm?searchValue={ip}&"
    response = requests.get(url, headers=headers)
    # 检查响应状态码
    if response.status_code != 200:
        return False, f"{task_name} 查询 {ip} 信息失败，状态码: {response.status_code}"

    # 解析网页内容
    soup = BeautifulSoup(response.text, 'html.parser')
    # 获取复选框
    table_list = soup.find_all('table')
    if len(table_list) < 5:
        return False, f"{task_name} 没有找到ip {ip} table标签"

    # 第四个table标签就是ip所在的table
    target_table = table_list[3]
    # 获取行
    all_rows = target_table.find_all('tr')
    if len(all_rows) < 1:
        print(f"{task_name} 没有找到ip {ip} tr标签\n", end='')
        return False, f"{task_name} 没有找到ip {ip} tr标签"

    # ip信息是唯一的，所以只取第一行
    row = all_rows[0]
    all_cols = row.find_all('td')
    remote_ip = all_cols[2].text.strip()
    if remote_ip != ip:
        return False, f"{task_name} 查询到的ip {remote_ip} 与请求的ip {ip} 不一致"
    # 状态
    status = all_cols[5].text.strip()
    # 使用类型
    useType = all_cols[6].text.strip()
    if status != '使用中':
        return True, {'ip': ip, 'status': status, 'useType': useType, 'deviceCodeUrl': None}
    # 所属设备编码
    deviceCodeUrl = all_cols[7].find('a').get('href')
    deviceCodeUrl = f"http://itsm.cnsuning.com{deviceCodeUrl}"
    return True, {'ip': ip, 'status': status, 'useType': useType, 'deviceCodeUrl': deviceCodeUrl}


def parse_server_page(link_info, device_code_url):
    """
    解析服务器server页面
    :param device_code_url:
    :param link_info:
    :return:
    """
    vip_info = link_info['vip_info']
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0',
        'Cookie': COOKIE_STR
    }
    task_name = str(link_info.get('taskName')).strip()
    vip = link_info.get('vip')
    # server页面中的具体server信息是通过页面server.js加载的，其实际的url是 loadServerListView.htm
    # 所以这里先从device_code_url中提取出所属设备编码，然后拼接成新的url
    device_code = device_code_url.split('=')[-1]
    url = f'http://itsm.cnsuning.com/traffic-web-in/server/loadServerListView.htm?searchValue={device_code}'
    response = requests.get(url, headers=headers)

    # 检查响应状态码
    if response.status_code != 200:
        print(f"{task_name} 请求失败，状态码: {response.status_code}\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"请求失败，状态码: {response.status_code}"})
        return

    # 解析网页内容
    soup = BeautifulSoup(response.text, 'html.parser')
    # loadServerListView 返回的table只有一个，所以这里直接取第一个
    table = soup.find('table')
    if table is None:
        print(f"{task_name} 没有找到table标签, 跳过\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"{task_name} 没有找到table标签"})
        return
    # 只有一行，所以直接取第一个
    row = table.find('tr')
    if row is None:
        print(f"{task_name} 没有找到tr标签, 跳过\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"{task_name} 没有找到tr标签"})
        return

    # 获取列
    all_cols = row.find_all('td')
    # 获取ip
    ip = all_cols[5].text.strip()
    if ip != vip:
        print(f"{task_name} 查询到的ip {ip} 与请求的ip {vip} 不一致\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"{task_name} 查询到的ip {ip} 与请求的ip {vip} 不一致"})
        return
    # 获取系统信息
    sys_info = all_cols[12].text.strip()
    # 从 易付宝账务核心(FAEPP) 中提取出 FAEPP
    sys_en = sys_info.split('(')[1].split(')')[0]
    sys_cn = sys_info.split('(')[0]
    role = ''
    # 获取当前IP对应的备份网络ip,server类型的直接可以从页面上取到
    bk_ip = all_cols[25].text.strip()
    vip_dict = {
        'vip': vip,
        'ip': ip,
        'role': role,
        'sys_en': sys_en,
        'sys_cn': sys_cn,
        'bk_ip': bk_ip,
    }
    vip_info.append(vip_dict)


def parse_mysql_ha_page(link_info, device_code_url):
    """
    解析mysql高可用vip页面
    :param link_info:
    :param device_code_url:
    :return:
    """
    vip_info = link_info['vip_info']
    # 构造请求头部和Cookie
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0',
        'Cookie': COOKIE_STR
    }
    task_name = str(link_info.get('taskName')).strip()
    vip = link_info.get('vip')
    url = device_code_url
    response = requests.get(url, headers=headers)

    # 检查响应状态码
    if response.status_code != 200:
        print(f"{task_name} 请求失败，状态码: {response.status_code}\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"请求失败，状态码: {response.status_code}"})
        return

    # 解析网页内容
    soup = BeautifulSoup(response.text, 'html.parser')
    # 获取复选框
    table_list = soup.find_all('table')
    if len(table_list) < 4:
        print(f"{task_name} 没有找到table标签, 跳过\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"{task_name} 没有找到table标签"})
        return

    # 第四个table标签就是系统和数据库ip所在的table
    target_table = table_list[3]
    # 获取行
    all_rows = target_table.find_all('tr')
    if len(all_rows) < 1:
        print(f"{task_name} 没有找到tr标签, 跳过\n", end='')
        vip_info.append({'vip': vip, 'err_msg': f"{task_name} 没有找到tr标签"})
        return

    for row in all_rows:
        # 获取列
        all_cols = row.find_all('td')
        # 获取ip
        ip = all_cols[3].text.strip()
        # 获取系统信息
        sys_info = all_cols[7].text.strip()
        # 从 易付宝账务核心(FAEPP) 中提取出 FAEPP
        sys_en = sys_info.split('(')[1].split(')')[0]
        sys_cn = sys_info.split('(')[0]
        role = all_cols[11].text.strip()
        # 获取当前IP对应的备份网络ip
        bk_ip = get_bk_ip(task_name, ip)
        vip_dict = {
            'vip': vip,
            'ip': ip,
            'role': role,
            'sys_en': sys_en,
            'sys_cn': sys_cn,
            'bk_ip': bk_ip,
        }
        vip_info.append(vip_dict)


def get_vip_info_0(link_info):
    """
    从itsm获取vip信息
    :param link_info: 链路信息
    :return:
    """
    task_name = str(link_info.get('taskName')).strip()

    # 发起GET请求获取页面内容
    vip = get_vip_from_task_name(task_name)
    link_info['vip'] = vip
    # 查询内网ip基本信息
    suc, ip_info = query_ip_info(task_name, vip)
    if not suc:
        print(f"{task_name} 查询vip信息失败: {ip_info}\n", end='')
        link_info['vip_info'] = [{'vip': vip, 'err_msg': ip_info}]
        return
    # 状态: 未使用 使用中 保留
    statue = ip_info.get('status')
    if statue != '使用中':
        link_info['vip_info'] = [{'vip': vip, 'err_msg': statue}]
        return
    # 使用类型: 高可用VIP 服务器Server
    use_type = ip_info.get('useType')
    deviceCodeUrl = ip_info.get('deviceCodeUrl')
    # 存放最终解析到的vip信息
    if use_type == '服务器Server':
        #  服务器Server的处理
        link_info['vip_info'] = []
        parse_server_page(link_info, deviceCodeUrl)
    elif use_type == '高可用VIP':
        # mysql高可用vip的处理
        link_info['vip_info'] = []
        parse_mysql_ha_page(link_info, deviceCodeUrl)
    else:
        link_info['vip_info'] = [{'vip': vip, 'err_msg': f'不合法的使用类型: {use_type}'}]
    print(f"{task_name} vip信息获取成功 {link_info['vip_info']} \n", end='')


def set_vip_info(link_list):
    """
    获取链路对应的tps信息
    :param link_list: 链路列表
    :return:
    """
    executor = ThreadPoolExecutor(max_workers=10)
    # map方法会将link_list中的每个元素传递给get_tps_info_0函数
    executor.map(get_vip_info_0, link_list)
    # 这里等待所有任务执行完毕再返回这个函数
    executor.shutdown()


def save_excel(link_list):
    """
    写入link_list到excel中
    :param link_list: 链路列表
    :return: 工作表
    """
    wb = Workbook()
    if os.path.exists(EXCEL_FILE):
        # 存在则使用已存在的文件
        wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_name = 'vip_info'
    if sheet_name in wb.sheetnames:
        # 存在则删除
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    print('创建sheet页: {}'.format(sheet_name))

    header = ['vip', 'ip', 'role', 'sys_en', 'sys_cn', 'bk_ip']
    ws.append(header)
    for link_dict in link_list:
        vip_info = link_dict.get('vip_info')
        if vip_info is None:
            continue
        for vip in vip_info:
            row = [vip.get(key) for key in header]
            ws.append(row)
    wb.save(EXCEL_FILE)
    return wb


def save_database(link_list):
    """
    保存到数据库
    :return:
    """
    # 连接到MySQL数据库
    conn = pymysql.connect(
        host="10.237.217.107",
        user="fabu",
        password="2d8VfEk6jFD6",
        database="snrs_sit"
    )

    # 创建游标对象
    cursor = conn.cursor()

    # 执行SQL查询
    truncate_sql = "TRUNCATE TABLE vip_from_itsm;"
    cursor.execute(truncate_sql)

    insert_sql = "INSERT INTO vip_from_itsm (vip, ip, role, sys_en, sys_cn, bk_ip, err_msg) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    for link_dict in link_list:
        vip_info = link_dict.get('vip_info')
        if vip_info is None:
            continue
        for vip in vip_info:
            cursor.execute(insert_sql, (
                vip.get('vip'), vip.get('ip'), vip.get('role'), vip.get('sys_en'), vip.get('sys_cn'), vip.get('bk_ip'),
                vip.get('err_msg')))
    conn.commit()

    # 关闭游标和数据库连接
    cursor.close()
    conn.close()


all_link = get_link_list()
if len(all_link) == 0:
    print('链路数据为空')
    sys.exit()
print('vip数量: {}'.format(len(all_link)))
print('开始获取vip信息...')
set_vip_info(all_link)
# save_excel(all_link)
save_database(all_link)
print('vip信息获取完成')
