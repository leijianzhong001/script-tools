# -*- coding:utf-8 -*-
import os
import traceback

import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
import re
from openpyxl.workbook import Workbook
import pymysql.cursors
import tkinter as tk
import sys

"""
导出所有的订阅链路到excel中
"""

COOKIE_STR = ''
URL_GET_TASK_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTaskInfoByKind'
URL_GET_TPS_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTpsInfo'
URL_GET_CONFIG_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getConfigInfo'
URL_GET_IP_INFO = 'http://itsm.cnsuning.com/traffic-web-in/api/getServerByIpInfo.htm'
URL_EXECUTE_SQL = 'http://rdrsmgr2.cnsuning.com/mgr/serverProxy/getSqlExecute'
PAGE_SIZE = 100
EXCEL_FILE = 'D:\\file\\202407\\订阅链路导出.xlsx'
vip_info = {}


def get_vip_from_task_name(task_name):
    """
    从taskName中获取vip
    :return:
    """
    # 找到上面字符串中包含的ip
    return re.findall(r"\d+.\d+.\d+.\d+", task_name)[0]


def execute_sql(sql_statement):
    """
    保存到数据库
    :return:
    """
    if sql_statement is None:
        return None

    # 连接到MySQL数据库
    conn = pymysql.connect(
        host="10.237.217.107",
        user="fabu",
        password="2d8VfEk6jFD6",
        database="snrs_sit"
    )

    # 创建游标对象
    cursor = conn.cursor()

    # 执行SQL查询
    cursor.execute(sql_statement)

    if 'select' in sql_statement:
        # 获取所有记录列表
        results = cursor.fetchall()
        return results
    if 'insert' in sql_statement or 'update' in sql_statement or 'delete' in sql_statement:
        conn.commit()
        # 返回影响的行数
        return cursor.rowcount

    # 关闭游标和数据库连接
    cursor.close()
    conn.close()


def init_vip_info():
    global vip_info
    rows = execute_sql('select * from vip_from_itsm;')
    # 字典推导为一个key为vip， 值为dict的字典
    for row in rows:
        vip = row[0]
        if vip_info.get(vip) is None:
            vip_info[vip] = []
        db_list = vip_info[vip]
        row_dict = {'vip': row[0], 'ip': row[1], 'role': row[2], 'sys_en': row[3], 'sys_cn': row[4], 'bk_ip': row[5], 'err_msg': row[6]}
        db_list.append(row_dict)
    print('vip信息初始化完成！')


def set_consistent(link_info):
    """
    从itsm查备份网络ip对应的系统名称，并比较是否和当前系统一致
    :param link_info:
    :return:
    """
    rdrs_bk_ip = link_info.get('bkIp').strip()
    rdrs_sys_en = link_info.get('systemName').strip()
    task_name = link_info.get('taskName').strip()
    vip = get_vip_from_task_name(task_name).strip()

    if vip not in vip_info:
        print(f'{task_name} vip信息不存在！\n', end='')
        link_info['consistent'] = False
        link_info['consistent_reason'] = 'vip信息不存在'
        return

    db_list = vip_info.get(vip)
    sys_en_list = [db.get('sys_en') for db in db_list]
    link_info['itsmSystem'] = ','.join(f'{a}' for a in sys_en_list)
    # 将bk_ip_list中的所有元素拼接成一个字符串
    bk_ip_list = [db.get('bk_ip') for db in db_list]
    link_info['itsmBkIp'] = ','.join(f'{a}' for a in bk_ip_list)
    link_info['consistent'] = True
    link_info['consistent_reason'] = ''
    for db in db_list:
        sys_en = db.get('sys_en')
        if sys_en is None:
            link_info['consistent'] = False
            link_info['consistent_reason'] = db.get('err_msg')
            return

        if rdrs_sys_en != sys_en:
            link_info['consistent'] = False
            link_info['consistent_reason'] = 'vip已不归属于当前系统'
            return

    db_ip = [db.get('ip') for db in db_list]
    if rdrs_bk_ip not in bk_ip_list and rdrs_bk_ip not in db_ip and rdrs_bk_ip != vip:
        link_info['consistent'] = False
        link_info['consistent_reason'] = '备份网络ip已变更'


def get_config_info_0(link_info):
    """
    获取链路的配置信息
    :param link_info: 链路信息
    :return:
    """
    try:
        task_name = link_info.get('taskName')
        param = {'taskName': task_name}
        response = requests.get(URL_GET_CONFIG_INFO, cookies=cookies, headers=headers, params=param)
        if response.status_code != 200:
            link_info['bkIp'] = 'NA'
            return
        response_dict = response.json()
        endpoint_code = response_dict.get('code')
        if str(endpoint_code) != '200':
            link_info['bkIp'] = 'NA'
            return

        database_hostname = response_dict.get('result').get('database.hostname')
        link_info['bkIp'] = database_hostname
        set_consistent(link_info)
    except Exception as e:
        link_info['consistent'] = False
        link_info['consistent_reason'] = f'发生异常{e}'
        traceback.print_exc()


def set_config_info(link_list):
    """
    获取链路对应的tps信息
    :param link_list: 链路列表
    :return:
    """
    executor = ThreadPoolExecutor(max_workers=1)
    # map方法会将link_list中的每个元素传递给get_tps_info_0函数
    executor.map(get_config_info_0, link_list)
    # 这里等待所有任务执行完毕再返回这个函数
    executor.shutdown()


def get_tps_info_0(link_info):
    """
    获取某条链路指定的tps信息
    :param link_info: 链路信息
    :return:
    """
    param = {'taskName': link_info.get('taskName')}
    response = requests.get(URL_GET_TPS_INFO, cookies=cookies, headers=headers, params=param)
    if response.status_code != 200:
        link_info['haveData'] = True
        link_info['tps'] = str(['9999' for _ in range(15)])
        return
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        link_info['haveData'] = True
        link_info['tps'] = str(['9999' for _ in range(15)])
        return
    # 只取最近15分钟这一项
    fifteen_tps_dict = response_dict.get('result')[2]
    # value ["7847", "5509", "6853", "5951", "5517", "4141", "3960", "4155", "4476", "5328", "6633", "7318",…]
    tps = fifteen_tps_dict.get('value')
    link_info['haveData'] = tps.count('0') != 15
    link_info['tps'] = str(tps)


def set_tps_info(link_list):
    """
    获取链路对应的tps信息
    :param link_list: 链路列表
    :return:
    """
    executor = ThreadPoolExecutor(max_workers=5)
    # map方法会将link_list中的每个元素传递给get_tps_info_0函数
    executor.map(get_tps_info_0, link_list)
    # 这里等待所有任务执行完毕再返回这个函数
    executor.shutdown()


def cookie_to_dict():
    """
    将字符串形式的cookie转为字典格式
    :return:
    """
    # 列表推导
    cookie_entries = [(entry.strip().split('=')[0], entry.strip().split('=')[1]) for entry in COOKIE_STR.split(';')]
    # 字典推导
    cookie_dict = {k: v for k, v in cookie_entries}
    return cookie_dict


cookies = {}
headers = {'Content-Type': 'application/json'}


def get_page_link_data(page, size, kind):
    """
    获取分页数据
    :param kind:
    :param page: 页号
    :param size: 每页数据条数
    :return: 该页数据
    """
    link_page_data_param = {'taskNo': '', 'kind': kind, 'taskName': '', 'status': '', 'system': '', 'page': page,
                            'size': size}
    # 在Cookie Version 0中规定空格、方括号、圆括号、等于号、逗号、双引号、斜杠、问号、@，冒号，分号等特殊符号都不能作为Cookie的内容。
    response = requests.get(URL_GET_TASK_INFO, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        return None
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        return None
    result = response_dict.get('result')
    data = result.get('data')
    if data is None or len(data) == 0:
        return None
    total = result.get('total')
    size = result.get('size')
    return data, total, size


def get_inc_link_list():
    """
    获取订阅链路列表
    :return:
    """
    link_list = []
    # 先探测一下数据总条数
    page_data = get_page_link_data(1, 10, '-INC-')
    if page_data is None:
        return link_list
    total = page_data[1]
    print()
    print('链路类型: {}'.format('订阅链路'))
    print('链路总量: {}'.format(total))
    print('每页数量: {}'.format(PAGE_SIZE))
    print('最大页数: {}'.format(int(total / PAGE_SIZE) + 2))
    print('开始获取页数据...')
    # 下面的代码和  for (int i = 1; i <= total / PAGE_SIZE + 1; i++)  是等价的
    for page_no in range(1, int(total / PAGE_SIZE) + 2):
        page_data = get_page_link_data(page_no, PAGE_SIZE, '-INC-')
        if page_data is None:
            continue
        page_link_list = page_data[0]
        # 设置tps信息
        # set_tps_info(page_link_list)
        # 设置链路的配置信息
        set_config_info(page_link_list)
        # 获取链路的任务号，系统总监
        link_list.extend(page_link_list)
        print('当前页号: {} 数据获取完成'.format(page_no))
    print('订阅链路所有页数据获取完成！')
    return link_list


def write_link_list(inc_link_list):
    """
    写入link_list到excel中
    :param inc_link_list: 订阅链路列表
    :return: 工作表
    """
    wb = Workbook()
    if os.path.exists(EXCEL_FILE):
        # 存在则使用已存在的文件
        wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_name = '订阅链路'
    if sheet_name in wb.sheetnames:
        # 存在则删除
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    print('创建sheet页: {}'.format(sheet_name))

    header = ['taskId', 'taskName', 'taskStatus', 'systemName', 'clusterName', 'workIp', 'startTime', 'maxBinlogFile',
              'nowBinlogFile', 'minBinlogFile', 'nowPosition', 'trace', 'delay', 'delayOfnum', 'haveData', 'tps',
              'bkIp', 'itsmSystem', 'itsmBkIp', 'consistent', 'consistent_reason']
    ws.append(header)
    for link_dict in inc_link_list:
        row = [link_dict.get(key) for key in header]
        ws.append(row)
    wb.save(EXCEL_FILE)
    print('写入 {} 条订阅链路数据到excel成功！'.format(len(inc_link_list)))


def get_cookie_from_gui():
    # 创建一个Tkinter根窗口
    root = tk.Tk()
    root.title("itsm cookie")
    root.geometry("800x400")  # 设置窗口大小

    # 在窗口中添加一个Label
    label = tk.Label(root, text="输入itsm cookie(要通过验证码以后的cookie):")
    label.pack(pady=10)

    # 添加一个具有自定义高度和宽度的多行输入框
    text = tk.Text(root, height=20, width=70)  # 设置高度为5行，宽度为40字符
    text.pack(pady=10)

    # 添加一个按钮
    button = tk.Button(root, text="提交", command=lambda: submit(text.get("1.0", tk.END), root))
    button.pack(pady=10)

    # 显示窗口
    root.mainloop()


def submit(value, root):
    global COOKIE_STR
    COOKIE_STR = value.strip()
    if COOKIE_STR is None or COOKIE_STR == '':
        print('cookie为空')
        sys.exit()
    global cookies
    cookies = cookie_to_dict()

    root.destroy()  # 关闭窗口


def main():
    get_cookie_from_gui()
    # 初始化vip信息
    init_vip_info()
    # 获取全量的链路列表
    inc_link_list = get_inc_link_list()
    # 写入链路列表到excel中
    write_link_list(inc_link_list)


if __name__ == '__main__':
    main()
