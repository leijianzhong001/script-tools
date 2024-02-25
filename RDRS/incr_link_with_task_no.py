# -*- coding:utf-8 -*-
import os
import json
import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
import re
from openpyxl.workbook import Workbook

COOKIE_STR = 'experimentation_subject_id=IjFjYjNlNTYzLWRiZTAtNDM3ZC05MzFlLTk4OGY3MjZkYmI1MSI%3D--642f87fac21ad96137b0ca2b46f781df1d1781bf; tradeMA=68; _snvd=1702028146357WryCRiG+g6C; idsLoginUserIdLastTime=20013921; isso_ld=true; isso_us=20013921; _snma=1%7C170702760811448759%7C1707027608114%7C1707028405030%7C1708158908824%7C3%7C2; scm-prd=siEEAEBCE88E33795E2B38E9FAA441BC9B; authId=siFCEAC28880BB945DA2817E03E86187F8; secureToken=456F7039F7A0CD63EEC2D7ECEE1B2A19; JSESSIONID=14cje5suhddu1mqrt734hm7x9'
URL_GET_TASK_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTaskInfoByKind'
URL_GET_TPS_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTpsInfo'
URL_GET_CONFIG_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getConfigInfo'
URL_GET_IP_INFO = 'http://itsm.cnsuning.com/traffic-web-in/api/getServerByIpInfo.htm'
URL_EXECUTE_SQL = 'http://rdrsmgr2.cnsuning.com/mgr/serverProxy/getSqlExecute'
PAGE_SIZE = 100
EXCEL_FILE = 'D:\\file\\202402\\all_links_export_with_task_no.xlsx'
system_dict = {}


def get_vip_from_task_name(task_name):
    """
    从taskName中获取vip
    :return:
    """
    # 找到上面字符串中包含的ip
    return re.findall(r"\d+.\d+.\d+.\d+", task_name)[0]


def get_config_info_0(link_info):
    """
    获取链路的配置信息
    :param link_info: 链路信息
    :return:
    """
    task_name = link_info.get('taskName')
    param = {'taskName': task_name}
    response = requests.get(URL_GET_CONFIG_INFO, cookies=cookies, headers=headers, params=param)
    if response.status_code != 200:
        link_info['bkIp'] = 'NA'
        return
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        link_info['bkIp'] = 'NA'
        return

    if 'FZ' in task_name:
        database_hostname = response_dict.get('result').get('connection.url')
        link_info['vip'] = database_hostname
    else:
        database_hostname = response_dict.get('result').get('database.hostname')
        link_info['bkIp'] = database_hostname


def get_system_tech_manager():
    url = 'http://snrsadmin.cnsuning.com/snrs/sql/safe'
    snrs_sql_headers = {'Content-Type': 'application/json'}
    post_data = {
        "secretKey": "KiEyYYMdIhseLsMXbAhCmFHMjLwQqXrXRSrQfhsmZiHCKluxwZpPYuMHtJQjTZpp",
        "data": 'SELECT sndSystemEhShortName,sndSystemChName,techMangerNumber,techMangerName from snrssysteminfo;'
    }
    json_data = json.dumps(post_data)
    response = requests.post(url, data=json_data, headers=snrs_sql_headers)
    if response.status_code != 200:
        # 抛出异常
        raise Exception('get_system_tech_manager failed')
    json_response = json.loads(response.text)
    csv_data = json_response['data']
    global system_dict
    for sys in csv_data:
        sndSystemEhShortName = sys[0]
        sndSystemChName = sys[1]
        techMangerNumber = sys[2]
        techMangerName = sys[3]
        system_dict[sndSystemEhShortName] = {'sndSystemEhShortName': sndSystemEhShortName,
                                             'sndSystemChName': sndSystemChName, 'techMangerNumber': techMangerNumber,
                                             'techMangerName': techMangerName}
    print('获取全量系统技术负责人成功: ', len(system_dict))


def get_task_no_0(link_info):
    """
    获取任务号信息
    :param link_info: 链路信息
    :return:
    """
    task_name = link_info.get('taskName')
    vip = get_vip_from_task_name(task_name)
    task_no_sql = f"select * from BASE_JOB_PARAM where RWVIP = '{vip}' and INIT = 'false'"
    link_page_data_param = {'ip': '10.103.109.183/rdrsmgrprd1:3306', 'sql': task_no_sql}
    response = requests.get(URL_EXECUTE_SQL, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        print(f'{task_name} 调用URL_EXECUTE_SQL失败: {response.status_code} \n', end='')
        return []
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        print(f'{task_name} 查询task_no失败: {endpoint_code}\n', end='')
        return []
    base_job_param_list = response_dict.get('result')
    if len(base_job_param_list) == 0:
        print(f'{task_name} base_job_param_list为空 \n', end='')
        return []

    # TASK_NO, CONNECTOR_NAME, RUNNING_STATUS, TASK_TYPE
    task_no_list = [{'taskNo': base_job_param.get('TASK_NO'),
                     'runningStatus': base_job_param.get('RUNNING_STATUS'),
                     'taskType': base_job_param.get('TASK_TYPE')} for base_job_param in base_job_param_list if
                    base_job_param.get('RUNNING_STATUS') != 'cancel' and str(base_job_param.get('TASK_NO')).startswith('DY')]
    if len(task_no_list) == 0:
        print(f'{task_name} task_no_list为空 \n', end='')
        return []
    print(f'{task_name} task_no_list 长度为{len(task_no_list)}！')

    # 填充系统信息
    for base_job_param in task_no_list:
        task_no = base_job_param.get('taskNo')
        system_info_sql = f"select * from INCRE_SUBSCRIBE_SDK where SUBSCRIPTION_NO='{task_no}' and DELETED!=1"
        link_page_data_param = {'ip': '10.103.109.183/rdrsmgrprd1:3306', 'sql': system_info_sql}
        response = requests.get(URL_EXECUTE_SQL, cookies=cookies, headers=headers, params=link_page_data_param)
        if response.status_code != 200:
            print(f'{task_name} 调用URL_EXECUTE_SQL失败2: {response.status_code} \n', end='')
            continue
        response_dict = response.json()
        endpoint_code = response_dict.get('code')
        if str(endpoint_code) != '200':
            print(f'{task_name} {task_no} 系统信息失败: {endpoint_code}\n', end='')
            continue
        result_list = response_dict.get('result')
        if len(result_list) == 0:
            print(f'{task_name} {task_no} result_list为空 \n', end='')
            continue

        # 只有一条，所以直接取第一个
        incre_subscribe_sdk = result_list[0]
        clientId = incre_subscribe_sdk['CLIENTID']
        base_job_param['clientId'] = clientId
        # 填充技术负责人信息
        system = system_dict[clientId]
        base_job_param['techMangerNumber'] = system['techMangerNumber']
        base_job_param['techMangerName'] = system['techMangerName']
        base_job_param['sndSystemChName'] = system['sndSystemChName']
    return task_no_list


def set_config_info(link_list):
    """
    获取链路对应的tps信息
    :param link_list: 链路列表
    :return:
    """
    executor = ThreadPoolExecutor(max_workers=5)
    # map方法会将link_list中的每个元素传递给get_tps_info_0函数
    executor.map(get_config_info_0, link_list)
    # 这里等待所有任务执行完毕再返回这个函数
    executor.shutdown()


def cookie_to_dict():
    """
    将字符串形式的cookie转为字典格式
    :return:
    """
    # 列表推导
    cookie_entries = [(entry.strip().split('=')[0], entry.strip().split('=')[1]) for entry in COOKIE_STR.split(';')]
    # 字典推导
    cookie_dict = {k: v for k, v in cookie_entries}
    return cookie_dict


cookies = cookie_to_dict()
headers = {'Content-Type': 'application/json'}


def get_page_link_data(page, size, kind):
    """
    获取分页数据
    :param kind:
    :param page: 页号
    :param size: 每页数据条数
    :return: 该页数据
    """
    link_page_data_param = {'taskNo': '', 'kind': kind, 'taskName': '', 'status': '', 'system': '', 'page': page,
                            'size': size}
    # 在Cookie Version 0中规定空格、方括号、圆括号、等于号、逗号、双引号、斜杠、问号、@，冒号，分号等特殊符号都不能作为Cookie的内容。
    response = requests.get(URL_GET_TASK_INFO, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        return None
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        return None
    result = response_dict.get('result')
    data = result.get('data')
    if data is None or len(data) == 0:
        return None
    total = result.get('total')
    size = result.get('size')
    return data, total, size


def incr_convert_task_no_list(link_list):
    """
    获取任务号列表
    :param link_list: 链路列表
    :return:
    """
    final_link_list = []
    for i, link_info in enumerate(link_list):
        # 当前链路对应的多条任务号
        task_no_list = get_task_no_0(link_info)
        for task in task_no_list:
            # 将link_info中的字段更新到task中
            task.update(link_info)
        # 将合并之后的task添加到final_link_list中
        final_link_list.extend(task_no_list)
        print(f'{i}/{len(link_list)}    {link_info.get("taskName")} done！')
    return final_link_list


def get_inc_link_list():
    """
    获取订阅链路列表
    :return:
    """
    link_list = []
    # 先探测一下数据总条数
    page_data = get_page_link_data(1, 10, '-INC-')
    if page_data is None:
        return link_list
    total = page_data[1]
    print()
    print('链路类型: {}'.format('订阅链路'))
    print('链路总量: {}'.format(total))
    print('每页数量: {}'.format(PAGE_SIZE))
    print('最大页数: {}'.format(int(total / PAGE_SIZE) + 2))
    print('开始获取页数据...')
    # 下面的代码和  for (int i = 1; i <= total / PAGE_SIZE + 1; i++)  是等价的
    for page_no in range(1, int(total / PAGE_SIZE) + 2):
        page_data = get_page_link_data(page_no, PAGE_SIZE, '-INC-')
        if page_data is None:
            continue
        page_link_list = page_data[0]
        # 设置链路的配置信息
        # set_config_info(page_link_list)
        link_list.extend(page_link_list)
        print('当前页号: {} 数据获取完成'.format(page_no))
    print('订阅链路所有页数据获取完成！')
    final_link_list = incr_convert_task_no_list(link_list)
    return final_link_list


def write_incr_link_list(wb, max_no, inc_link_list):
    """
    :param wb: 工作簿
    :param max_no: sheet页最大序号，用于创建新的sheet页
    :param inc_link_list: 订阅链路列表
    写入订阅链路列表
    :return:
    """
    inc_sheet_name = "{}-{}".format('订阅链路', max_no)
    ws = wb.create_sheet(inc_sheet_name)
    print('创建sheet页: {}'.format(inc_sheet_name))

    header = ['taskNo', 'runningStatus', 'taskType', 'clientId', 'sndSystemChName', 'techMangerNumber',
              'techMangerName', 'taskId', 'taskName', 'taskStatus', 'systemName',
              'clusterName', 'workIp', 'startTime', 'maxBinlogFile',
              'nowBinlogFile', 'minBinlogFile', 'nowPosition', 'trace', 'delay', 'delayOfnum', 'haveData', 'tps',
              'bkIp', 'realSystemName', 'consistent']
    ws.append(header)
    for link_dict in inc_link_list:
        row = [link_dict.get(key) for key in header]
        ws.append(row)
    wb.save(EXCEL_FILE)
    print('写入 {} 条订阅链路数据到excel成功！'.format(len(inc_link_list)))


def main():
    # 获取全量系统技术负责人
    get_system_tech_manager()
    # 获取全量的链路列表
    inc_link_list = get_inc_link_list()
    # 写入链路列表到excel中
    wb = Workbook()
    if os.path.exists(EXCEL_FILE):
        # 存在则使用已存在的文件
        wb = openpyxl.load_workbook(EXCEL_FILE)
    # 获取全局最大序号
    max_no = 0
    if len(wb.sheetnames) != 0:
        max_no = max([int(my_sheet_names.split('-')[-1]) for my_sheet_names in wb.sheetnames])
    max_no = max_no + 1
    # 写入订阅链路列表
    write_incr_link_list(wb, max_no, inc_link_list)


if __name__ == '__main__':
    main()
