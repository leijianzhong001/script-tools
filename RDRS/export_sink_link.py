# -*- coding:utf-8 -*-
import os
import traceback

import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
import re
from openpyxl.workbook import Workbook
import pymysql.cursors

"""
导出所有的订阅链路到excel中
"""

COOKIE_STR = 'experimentation_subject_id=IjFjYjNlNTYzLWRiZTAtNDM3ZC05MzFlLTk4OGY3MjZkYmI1MSI%3D--642f87fac21ad96137b0ca2b46f781df1d1781bf; tradeMA=68; idsLoginUserIdLastTime=20013921; isso_ld=true; isso_us=20013921; scm-prd=siC07E171F39F366D7AD3261D61FEAE5EA; custno=20013921; _snma=1%7C170702760811448759%7C1707027608114%7C1708581444417%7C1708584359850%7C7%7C4; _snmp=17085843597737784; _snmb=170858435985351870%7C1708584359863%7C1708584359854%7C1; _snvd=1708581338362qBSqLmb+EpD; secureToken=C4CC3AC061BB215FB7DE014850E1E3B0; JSESSIONID=u3bq9cm3wf9f1nzsxqrppl1py; authId=siC53952DE4A5C5DC5674623A7E58965BE'
URL_GET_TASK_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTaskInfoByKind'
URL_GET_TPS_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getTpsInfo'
URL_GET_CONFIG_INFO = 'http://rdrsmgr2.cnsuning.com/mgr/taskInfo/getConfigInfo'
URL_GET_IP_INFO = 'http://itsm.cnsuning.com/traffic-web-in/api/getServerByIpInfo.htm'
URL_EXECUTE_SQL = 'http://rdrsmgr2.cnsuning.com/mgr/serverProxy/getSqlExecute'
PAGE_SIZE = 100
EXCEL_FILE = 'D:\\file\\202402\\同步链路导出.xlsx'
vip_info = {}


def get_vip_from_task_name(task_name):
    """
    从taskName中获取vip
    :return:
    """
    # 找到上面字符串中包含的ip
    return re.findall(r"\d+.\d+.\d+.\d+", task_name)[0]


def execute_sql(sql_statement):
    """
    保存到数据库
    :return:
    """
    if sql_statement is None:
        return None

    # 连接到MySQL数据库
    conn = pymysql.connect(
        host="10.237.217.107",
        user="fabu",
        password="2d8VfEk6jFD6",
        database="snrs_sit"
    )

    # 创建游标对象
    cursor = conn.cursor()

    # 执行SQL查询
    cursor.execute(sql_statement)

    if 'select' in sql_statement:
        # 获取所有记录列表
        results = cursor.fetchall()
        return results
    if 'insert' in sql_statement or 'update' in sql_statement or 'delete' in sql_statement:
        conn.commit()
        # 返回影响的行数
        return cursor.rowcount

    # 关闭游标和数据库连接
    cursor.close()
    conn.close()


def cookie_to_dict():
    """
    将字符串形式的cookie转为字典格式
    :return:
    """
    # 列表推导
    cookie_entries = [(entry.strip().split('=')[0], entry.strip().split('=')[1]) for entry in COOKIE_STR.split(';')]
    # 字典推导
    cookie_dict = {k: v for k, v in cookie_entries}
    return cookie_dict


cookies = cookie_to_dict()
headers = {'Content-Type': 'application/json'}


def get_page_link_data(page, size, kind):
    """
    获取分页数据
    :param kind:
    :param page: 页号
    :param size: 每页数据条数
    :return: 该页数据
    """
    link_page_data_param = {'taskNo': '', 'kind': kind, 'taskName': '', 'status': '', 'system': '', 'page': page,
                            'size': size}
    # 在Cookie Version 0中规定空格、方括号、圆括号、等于号、逗号、双引号、斜杠、问号、@，冒号，分号等特殊符号都不能作为Cookie的内容。
    response = requests.get(URL_GET_TASK_INFO, cookies=cookies, headers=headers, params=link_page_data_param)
    if response.status_code != 200:
        return None
    response_dict = response.json()
    endpoint_code = response_dict.get('code')
    if str(endpoint_code) != '200':
        return None
    result = response_dict.get('result')
    data = result.get('data')
    if data is None or len(data) == 0:
        return None
    total = result.get('total')
    size = result.get('size')
    return data, total, size


def get_sink_link_list():
    """
    获取同步链路列表
    :return:
    """
    link_list = []
    # 先探测一下数据总条数
    page_data = get_page_link_data(1, 10, '-SINK-')
    if page_data is None:
        return link_list
    total = page_data[1]
    print()
    print('链路类型: {}'.format('同步链路'))
    print('链路总量: {}'.format(total))
    print('每页数量: {}'.format(PAGE_SIZE))
    print('最大页数: {}'.format(int(total / PAGE_SIZE) + 2))
    print('开始获取页数据...')
    # 下面的代码和  for (int i = 1; i <= total / PAGE_SIZE + 1; i++)  是等价的
    for page_no in range(1, int(total / PAGE_SIZE) + 2):
        page_data = get_page_link_data(page_no, PAGE_SIZE, '-SINK-')
        if page_data is None:
            continue
        page_link_list = page_data[0]
        link_list.extend(page_link_list)
        print('当前页号: {} 数据获取完成'.format(page_no))
    print('同步链路所有页数据获取完成！')
    return link_list


def write_link_list(inc_link_list):
    """
    写入link_list到excel中
    :param inc_link_list: 订阅链路列表
    :return: 工作表
    """
    wb = Workbook()
    if os.path.exists(EXCEL_FILE):
        # 存在则使用已存在的文件
        wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet_name = '同步链路'
    if sheet_name in wb.sheetnames:
        # 存在则删除
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    print('创建sheet页: {}'.format(sheet_name))

    header = ['taskId', 'taskName', 'taskStatus', 'systemName', 'clusterName', 'workIp', 'startTime', 'maxBinlogFile',
              'nowBinlogFile', 'minBinlogFile', 'nowPosition', 'trace', 'delay', 'delayOfnum', 'haveData', 'tps']
    ws.append(header)
    for link_dict in inc_link_list:
        row = [link_dict.get(key) for key in header]
        ws.append(row)
    wb.save(EXCEL_FILE)
    print('写入 {} 条同步链路数据到excel成功！'.format(len(inc_link_list)))


def main():
    # 获取全量的链路列表
    inc_link_list = get_sink_link_list()
    # 写入链路列表到excel中
    write_link_list(inc_link_list)


if __name__ == '__main__':
    main()
